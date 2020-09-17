import xlrd
import pandas as pd
from openpyxl import load_workbook
import time
import os
import numpy as np
import warnings
import msoffcrypto
import datetime
from datetime import timedelta
import decimal
warnings.simplefilter(action='ignore', category=Warning)
OutputFolder = os.path.join(os.getcwd(), 'Output Folder')
while True:
    dt = datetime.datetime.utcnow()+timedelta(hours=-8)
    print('Current time is {}.'.format(dt),end="\r", flush=True)
    time.sleep(1)
    # if dt.hour == 2:
    if True:
        print('run')
        file = msoffcrypto.OfficeFile(open("IL.LYS.2261 PS.xlsx", "rb"))
        file.load_key(password="erh1")
        file.decrypt(open("decrypted.xlsx", "wb"))
        print('\n-- Spreadsheet decrypted --')
        outputFilePath = os.path.join(OutputFolder, 'Power Data Log.xlsx') #1 - 45
        outputFilePath1 = os.path.join(OutputFolder, 'Mass Removal.xlsx') # 46 - 76
        outputFilePath2 = os.path.join(OutputFolder, 'Condenser And Drip.xlsx') # 77 - 108
        outputFilePath3 = os.path.join(OutputFolder, 'Temps.xlsx') # 109 - 172
        inputFilePath = '{}/decrypted.xlsx'.format(os.path.dirname(os.path.abspath(__file__)))
        writer = pd.ExcelWriter(outputFilePath, engine='openpyxl')
        writer1 = pd.ExcelWriter(outputFilePath1, engine='openpyxl')
        writer2 = pd.ExcelWriter(outputFilePath2, engine='openpyxl')
        writer3 = pd.ExcelWriter(outputFilePath3, engine='openpyxl')
        a = 0
        paw = []
        pae = []
        par = []
        pat = []
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        writer1.sheets = dict((wsa.title, wsa) for wsa in writer1.book.worksheets)
        writer2.sheets = dict((wsb.title, wsb) for wsb in writer2.book.worksheets)
        writer3.sheets = dict((wsc.title, wsc) for wsc in writer3.book.worksheets)
        wb = load_workbook(inputFilePath, data_only=True)
        sh = wb.worksheets[1]
        noRows = sh.max_row
        noCol = sh.max_column+1

        def getColVal(row, col, a= a):
            col = sh.cell(row,col).value
            if col == None:
                col = ''
            return col

        def writeExcel(dataFrame):
            df = dataFrame

            df.to_excel(writer,index=False,header=False)
            writer.close()

        def writeExcel1(dataFrame):
            df = dataFrame

            df.to_excel(writer1,index=False,header=False)
            writer1.close()

        def writeExcel2(dataFrame):
            df = dataFrame

            df.to_excel(writer2,index=False,header=False)
            writer2.close()

        def writeExcel3(dataFrame):
            df = dataFrame

            df.to_excel(writer3,index=False,header=False)
            writer3.close()

        for p in range(0,noRows):
            paw.append([])
            pae.append([])
            par.append([])
            pat.append([])
            
        for i in range(2, noRows):
            for j in range(1, noCol):
                if j < 15:
                    try:
                        if j >= 8 and j <= 12:
                            try:
                                paw[i].append(int(decimal.Decimal(getColVal(i,j)*100).quantize(decimal.Decimal('1'), rounding=decimal.ROUND_HALF_UP)))
                            except:
                                paw[i].append(getColVal(i,j))
                        else:
                            paw[i].append(round(getColVal(i,j)))
                    except:
                        paw[i].append(getColVal(i,j))
                elif j >= 46 and j <= 76:
                    if j >= 51 and j <= 52:
                        try:
                            pae[i].append(round(getColVal(i,j)))
                        except:
                            pae[i].append((getColVal(i,j)))
                    elif j >= 53 and j <= 54:
                        try:
                            pae[i].append(round(getColVal(i,j), 13))
                        except:
                            pae[i].append((getColVal(i,j)))
                    else:
                        try:
                            pae[i].append(np.round(getColVal(i,j), 1))
                        except:
                            pae[i].append(getColVal(i,j))
                elif j >= 77 and j<= 108:
                    if j >= 82 and j <= 83:
                        try:
                            par[i].append(round(getColVal(i,j)))
                        except:
                            par[i].append((getColVal(i,j)))
                    else:
                        try:
                            par[i].append(np.round(getColVal(i,j), 1))
                        except:
                            par[i].append(getColVal(i,j))
                elif j >= 109 and j <= 172:
                    if j >= 111 and j <= 117:
                        try:
                            pat[i].append(round(getColVal(i,j)))
                        except:
                            pat[i].append((getColVal(i,j)))
                    else:
                        try:
                            pat[i].append(np.round(getColVal(i,j), 1))
                        except:
                            pat[i].append(getColVal(i,j))
        
        df = pd.DataFrame(paw)
        de = pd.DataFrame(pae)
        dr = pd.DataFrame(par)
        dt = pd.DataFrame(pat)

        for i in range(7,12):
            for j in range(1,noRows):
                if isinstance(df[i][j], int):
                    try:
                        df[i][j] = str(df[i][j]) + '%'
                    except:
                        print("something wrong with df")

        for i in range(1, noRows):
            for j in range(1, noCol):
                try:
                    df[i][j] = df[i][j].strftime('%m/%d/%Y')
                except:
                    pass

        for i in range(1, noRows):
            for j in range(1, noCol):
                try:
                    de[i][j] = de[i][j].strftime('%m/%d/%Y')
                except:
                    pass

        for i in range(1, noRows):
            for j in range(1, noCol):
                try:
                    dr[i][j] = dr[i][j].strftime('%m/%d/%Y')
                except:
                    pass

        for i in range(1, noRows):
            for j in range(1, noCol):
                try:
                    dt[i][j] = dt[i][j].strftime('%m/%d/%Y')
                except:
                    pass
        df = df.drop([df.index[0], df.index[1]])
        de = de.drop([de.index[0], de.index[1]])
        dr = dr.drop([dr.index[0], dr.index[1]])
        dt = dt.drop([dt.index[0], dt.index[1]])
        df.pop(0)
        de.pop(0)
        dr.pop(0)
        dt.pop(0)
        print(df)
        print(de)
        print(dr)
        print(dt)
        writeExcel(df)
        writeExcel1(de)
        writeExcel2(dr)
        writeExcel3(dt)
        try:
            os.system('rm decrypted.xlsx')
        except:
            pass

        print('Export Finished!')
        time.sleep(7200)
