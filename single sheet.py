import xlrd
import pandas as pd
from openpyxl import load_workbook
import time
import os
import numpy as np
import warnings
import msoffcrypto
from datetime import datetime
import decimal

warnings.simplefilter(action='ignore', category=Warning)

file = msoffcrypto.OfficeFile(open("IL.LYS.2261 PS.xlsx", "rb"))
file.load_key(password="erh1")
file.decrypt(open("decrypted.xlsx", "wb"))
print('-- Spreadsheet decrypted --')
outputFilePath = "./Power Data Log.xlsx" #1 - 45
outputFilePath1 = "./Mass Removal.xlsx" # 46 - 76
outputFilePath2 = "./Condenser And Drip.xlsx" # 77 - 108
outputFilePath3 = "./Temps.xlsx" # 109 - 172
inputFilePath = '{}/decrypted.xlsx'.format(os.path.dirname(os.path.abspath(__file__)))
writer = pd.ExcelWriter(outputFilePath, engine='openpyxl')
writer = pd.ExcelWriter(outputFilePath1, engine='openpyxl')
writer = pd.ExcelWriter(outputFilePath2, engine='openpyxl')
writer = pd.ExcelWriter(outputFilePath3, engine='openpyxl')
a = 0
paw = []
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
wb = load_workbook(inputFilePath, data_only=True)
sh = wb.worksheets[1]
noRows = sh.max_row
noCol = sh.max_column+1

def getColVal(row, col, a= a):
    col = sh.cell(row,col).value
    if col == None:
       col = ''
    return col

def writeExcel(dataFrame):
    df = dataFrame

    df.to_excel(writer,index=False,header=False)
    writer.close()



for p in range(noRows):
    paw.append([])
    
for i in range(1, noRows):
    for j in range(1, noCol):
        if j < 15:
            try:
                if j >= 8 and j <= 12:
                    try:
                        paw[i].append(int(decimal.Decimal(getColVal(i,j)*100).quantize(decimal.Decimal('1'), rounding=decimal.ROUND_HALF_UP)))
                    except:
                        paw[i].append(getColVal(i,j))
                else:
                    paw[i].append(round(getColVal(i,j)))
            except:
                paw[i].append(getColVal(i,j))
        elif j >= 51 and j <= 52:
            try:
                paw[i].append(round(getColVal(i,j)))
            except:
                paw[i].append((getColVal(i,j)))
        elif j >= 53 and j <= 54:
            try:
                paw[i].append(round(getColVal(i,j), 13))
            except:
                paw[i].append((getColVal(i,j)))
        elif j >= 82 and j <= 83:
            try:
                paw[i].append(round(getColVal(i,j)))
            except:
                paw[i].append((getColVal(i,j)))
        elif j >= 111 and j <= 117:
            try:
                paw[i].append(round(getColVal(i,j)))
            except:
                paw[i].append((getColVal(i,j)))
        else:
            try:
                paw[i].append(np.round(getColVal(i,j), 1))
            except:
                paw[i].append(getColVal(i,j))

df = pd.DataFrame(paw)

for i in range(7,12):
    for j in range(noRows):
        if isinstance(df[i][j], int):
            try:
                df[i][j] = str(df[i][j]) + '%'
            except:
                pass

for i in range(1, noRows):
    for j in range(1, noCol):
        try:
            df[i][j] = df[i][j].strftime('%m/%d/%Y')
        except:
            pass
print(df)
writeExcel(df)
try:
    os.system('rm decrypted.xlsx')
except:
    pass

print('Export Finished!')
